from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
hdr_fill = PatternFill('solid', fgColor='1A3A5C')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=10, color='1A1A2E')
cell_align = Alignment(vertical='top', wrap_text=True)
cell_align_c = Alignment(horizontal='center', vertical='top', wrap_text=True)
stripe_fill = PatternFill('solid', fgColor='EDF2F7')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
title_font = Font(name='Arial Black', bold=True, size=12, color='1A3A5C')

def style_sheet(ws, col_widths, title_row=0):
    """Apply standard styling to a sheet"""
    if title_row > 0:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_widths))
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30
        data_start = title_row + 1
    else:
        data_start = 1

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row
    for c in range(1, len(col_widths) + 1):
        cell = ws.cell(row=data_start, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border
    ws.row_dimensions[data_start].height = 30

    # Data rows
    for r in range(data_start + 1, ws.max_row + 1):
        for c in range(1, len(col_widths) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = cell_font
            cell.alignment = cell_align_c if c == 1 else cell_align
            cell.border = thin_border
            if (r - data_start) % 2 == 0:
                cell.fill = stripe_fill
        ws.row_dimensions[r].height = 40

def write_table(ws, headers, data, start_row=1):
    """Write headers and data rows, return next empty row"""
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    for r, row in enumerate(data, start_row + 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    return start_row + len(data) + 1

# ============================================
# Sheet 1: 功能等级总览
# ============================================
ws1 = wb.active
ws1.title = "功能等级总览"
ws1.cell(row=1, column=1, value="AIS 31 功能等级总览").font = title_font
ws1.merge_cells('A1:I1')
cols1 = [12, 8, 12, 10, 18, 14, 22, 28, 18]
headers1 = ['类别', '等级', '类型', '安全强度', '后处理要求', '随机模型', '在线测试', '适用场景', '熵要求']
data1 = [
    ['DRNG (确定性)', 'DRG.2', '确定性', '基础级', 'ϕ或ψ是密码学的', '不需要', '不需要', '不需要保证新鲜熵的密码应用\n(一般密钥、Nonces、挑战值)', '种子: min-entropy≥240 bits\n或 Shannon熵≥250 bits'],
    ['DRNG (确定性)', 'DRG.3', '确定性', '增强级', 'ϕ和ψ都必须是密码学的\nϕ必须是单向函数', '不需要', '不需要', '除需要新鲜熵外的所有密码应用\n(证书签名、长期密钥)', '同DRG.2 + 支持种子树\n(子DRNG可由父DRNG种子)'],
    ['DRNG (确定性)', 'DRG.4', '混合型(Hybrid)', 'DRNG最高级', '密码学后处理(强制)\n必须带记忆', '不需要', '需要(熵源部分)', '最高安全需求的DRNG应用\n(需持续熵刷新的场景)', '同DRG.3 + 持续熵刷新\n熵源: PTG.2/PTG.3/NTG.1'],
    ['PTRNG (物理真随机)', 'PTG.2', '物理真随机', '基础级', '可选: 非密码学后处理\n(Von Neumann, XOR, 压缩等)', '强制要求\n(核心评估依据)', '强制要求\n(在线+总失效+启动)', 'DRNG种子源、PTG.3核心组件\n不推荐直接用于任意密码应用', 'Shannon熵每比特>0.9998\n或 min-entropy每比特>0.98'],
    ['PTRNG (物理真随机)', 'PTG.3', '物理真随机(混合)', '整体最强级', '密码学后处理(强制)\n必须带记忆, c_rate≥1', '强制要求\n(对原始随机数)', '强制要求\n(对噪声源部分)', '最高安全需求应用\n(长期密钥、高安全签名等)', '同PTG.2 + 支持多种熵声明\n(Shannon/min-entropy/特定熵界)'],
    ['NPTRNG (非物理真随机)', 'NTG.1', '非物理真随机', '—', '可选', '不要求\n(无法建模)', '强制要求\n(在线+总失效)', 'DRNG种子源\n跨平台熵收集场景', '仅使用min-entropy\n(保守估计,每比特熵下界)'],
]
write_table(ws1, headers1, data1, 2)
style_sheet(ws1, cols1, 1)

# ============================================
# Sheet 2: DRNG详细要求对比
# ============================================
ws2 = wb.create_sheet("DRNG详细要求对比")
ws2.cell(row=1, column=1, value="DRNG 功能等级详细要求对比 (DRG.2 / DRG.3 / DRG.4)").font = title_font
ws2.merge_cells('A1:F1')
cols2 = [24, 18, 18, 18, 22, 28]
headers2 = ['要求项', 'DRG.2', 'DRG.3', 'DRG.4', '分类', '说明']
data2 = [
    ['种子来源', 'TRNG (推荐PTG.2/3或NTG.1)\n允许使用未认证TRNG但需额外验证', 'TRNG (同DRG.2) +\n允许DRNG种子DRNG(种子树)', 'TRNG or DRNG\n+ 持续熵输入', '安全要求', 'DRG.3新增种子树机制，减轻多实例认证负担'],
    ['请求限制', '最多2^48次请求\n单次≤2^19 bits', '同DRG.2', '同DRG.3', '安全要求', '请求间需重新种子或重新播种'],
    ['有效内部状态', '≥252 bits\n(在攻击者已知大量内部随机数的假设下)', '同DRG.2', '≥252 bits\n+ 密码学后处理内部状态', '安全要求', '内部状态大小直接决定暴力破解难度'],
    ['初始熵要求', 'min-entropy≥240 bits\n或Shannon熵≥250 bits\n(Shannon需平稳分布TRNG)', '同DRG.2', '同DRG.3\n+ 持续熵输入保证', '熵要求', '使用已认证TRNG可大幅简化验证'],
    ['前向保密', '✓ (内部随机数粒度)', '同DRG.2', '同DRG.3', '算法要求', '已知当前内部状态→不可计算之前的内部随机数'],
    ['后向保密', '✓ (内部随机数粒度)', '同DRG.2', '同DRG.3', '算法要求', '已知当前内部状态→不可计算之后的内部随机数'],
    ['增强后向保密', '— (不要求)', '✓ (请求粒度)', '同DRG.3', '算法要求', '即使已知某个请求的全部输出，也不能计算后续请求的输出\nDRG.3相比DRG.2的核心增强'],
    ['状态转移函数ϕ', 'ϕ或ψ之一是密码学的', 'ϕ必须是密码学的\n且是单向函数(One-Way)', 'ϕ必须是密码学的\n且是单向函数(One-Way)', '算法要求', 'DRG.3/4要求更严格：两个函数都必须是密码学的'],
    ['输出函数ψ', 'ϕ或ψ之一是密码学的', 'ψ必须是密码学的', '同DRG.3', '算法要求', 'DRG.3/4两个函数都必须是密码学的'],
    ['额外输入安全性', '即使攻击者控制额外输入\n也不削弱DRNG强度', '同DRG.2', '同DRG.3', '安全要求', '额外输入可能来自不可信源'],
    ['统计不可区分性', '有强证据(理论/统计)表明\n内部随机数与理想RNG不可区分', '同DRG.2', '同DRG.3', '算法要求', '可基于理论分析或统计测试'],
    ['种子树结构', '— (不支持)', '✓ 支持DRNG种子传递\n根DRNG必须用TRNG种子\n高度有限制', '✓ 支持(同DRG.3)', '结构特性', '根DRNG(TRNG)→子DRNG→孙DRNG\n逐级种子传递'],
    ['密码学后处理', '— (不要求单独后处理)', '— (ϕ和ψ已满足)', '✓ 强制，必须带记忆\n持续由熵源刷新', 'DRG.4特有', '这是DRG.4区别于DRG.3的核心特征'],
    ['持续熵刷新', '— (纯确定性)', '— (纯确定性)', '✓ 输出由熵输入持续刷新', 'DRG.4特有', '防止长时间运行导致的熵耗尽'],
    ['混合型(Hybrid)', '不是', '不是', '是 (Hybrid RNG)', '分类', '结合了DRNG的计算安全性和TRNG的信息论安全性'],
]
write_table(ws2, headers2, data2, 2)
style_sheet(ws2, cols2, 1)

# ============================================
# Sheet 3: PTRNG详细要求对比
# ============================================
ws3 = wb.create_sheet("PTRNG详细要求对比")
ws3.cell(row=1, column=1, value="PTRNG 功能等级详细要求对比 (PTG.2 / PTG.3)").font = title_font
ws3.merge_cells('A1:E1')
cols3 = [24, 18, 20, 20, 28]
headers3 = ['要求项', 'PTG.2', 'PTG.3', '分类', '说明']
data3 = [
    ['随机模型\n(Stochastic Model)', '强制要求\n原始随机数的随机行为\n必须追溯到物理现象', '强制要求(同PTG.2)\n对原始随机数部分', '核心要求', 'PTRNG评估的核心。不接受纯黑盒测试\n替代随机模型分析'],
    ['时间局部平稳性\n(Time-Local Stationarity)', '强制要求\n原始随机数满足', '强制要求(同PTG.2)', '统计要求', '确保不同时间段的随机数特性一致'],
    ['在线测试\n(Online Test)', '强制要求\n基于随机模型验证有效性\n实时监控噪声源质量', '强制要求(同PTG.2)', '测试要求', '低误报率(A_good区域)\n高检测率(A_bad区域)'],
    ['总失效测试\n(Total Failure Test)', '强制要求\n检测噪声源完全失效\n几乎立即触发', '强制要求(同PTG.2)', '测试要求', '检测硬件故障导致的噪声源完全失效'],
    ['启动测试\n(Start-up Test)', '强制要求\n上电/复位后执行\n验证噪声源初始正常', '强制要求(同PTG.2)', '测试要求', '在输出任何随机数之前必须通过'],
    ['Shannon熵/比特', '>0.9998 per bit\n(如果声明Shannon熵)', '>0.9998 per bit\n或声明PTRNG特定熵界', '熵要求', '需要平稳分布假设\nPTG.3更灵活: 可声明特定熵界'],
    ['Min-entropy/比特', '>0.98 per bit\n(如果声明min-entropy)', '>0.98 per bit\n或声明PTRNG特定熵界', '熵要求', '最保守的熵度量\nPTG.3可声明更精确的特定熵界'],
    ['密码学后处理\n(Cryptographic Post-processing)', '可选(可不使用)\n如使用: 可选带记忆或不带记忆', '强制使用\n必须带记忆(有内部状态)\nc_rate = n/m ≥ 1', 'PTG.3核心要求', 'PTG.3的核心区别:\n密码学后处理不是可选项'],
    ['压缩率 c_rate', '不适用\n(如使用压缩后处理则参考)', 'c_rate = n/m ≥ 1\n(输入比特/输出比特)', 'PTG.3特有', 'c_rate >1才能增加每比特熵\nc_rate <1违反PTG.3要求'],
    ['后处理类型', '算法后处理(非密码学):\n- Von Neumann去偏\n- XOR压缩\n- 固定压缩率\n- 恒等映射(无后处理)', '密码学后处理(强制):\n- Hash DRBG\n- AES-OFB\n- SHA-256 based', '后处理', '非密码学 vs 密码学的本质区别:\n后处理是否有内部状态/记忆'],
    ['内部随机数增强要求', '无特殊增强要求\n(原始随机数等于内部随机数\n或仅经算法后处理)', '中间随机数→密码学后处理\n→内部随机数\n内部随机数满足DRG.3级安全性', 'PTG.3特有', '即使PTG.2核心暂时失效\nPTG.3整体仍保持DRG.3安全性'],
    ['多种熵声明方式', '二选一:\nShannon或min-entropy', '多种选项:\n① Shannon >0.9998\n② min-entropy >0.98\n③ PTRNG特定熵界', 'PTG.3特有', '选项③允许基于随机模型计算\n更精确的熵界(更灵活)'],
    ['DRNG后备安全层', '无', '有: 密码学后处理即使自主运行\n也满足DRG.3要求', 'PTG.3特有', '额外安全层:\nPTG.2核心失效→DRG.3后备'],
    ['典型设计', '物理噪声源→数字化\n→(可选)算法后处理\n→内部随机数', 'PTG.2核心→中间随机数\n→DRG.3密码学后处理\n→内部随机数', '设计模式', 'PTG.3 = PTG.2核心 + DRG.3后处理'],
]
write_table(ws3, headers3, data3, 2)
style_sheet(ws3, cols3, 1)

# ============================================
# Sheet 4: 熵度量对比
# ============================================
ws4 = wb.create_sheet("熵度量对比")
ws4.cell(row=1, column=1, value="熵度量对比 (Entropy Measures Comparison)").font = title_font
ws4.merge_cells('A1:H1')
cols4 = [16, 22, 12, 20, 18, 18, 22, 18]
headers4 = ['熵类型', '公式', '保守程度', '优点', '缺点', '适用场景', 'AIS31使用位置', '与GuessWork关系']
data4 = [
    ['Min-Entropy\n(最小熵)', 'H_min = -log₂(max(pᵢ))', '最保守\n(★★★★★)', '最可靠的安全度量\n不依赖于分布假设\n计算简单', '可能低估实际熵\n对均匀分布以外的分布\n过于悲观', '安全认证 (首选)\nNPTRNG (唯一选项)\n保守熵下界估计', 'DRG.2/3/4: 初始内部状态\nPTG.2/3: 每比特熵声明\nNTG.1: 唯一使用的熵度量', 'H_min ≤ w_λ(X) ≤ H_min - log₂(1-λ)\n直接用于评估猜测难度'],
    ['Shannon熵\n(香农熵)', 'H = -Σ pᵢ·log₂(pᵢ)', '中等\n(★★★☆☆)', '信息论标准度量\n反映平均不确定性\n被广泛理解', '需要平稳分布假设\n不直接对应安全强度\n可能高估实际安全性', 'PTRNG (需随机模型)\nPTG.2/3允许声明\n平稳分布场景', 'DRG.2/3/4: 初始内部状态(备选)\nPTG.2/3: 每比特熵声明\n(需PTRNG随机模型)', '无直接关系\n可能严重低估猜测难度'],
    ['Collision熵\n(碰撞熵)', 'H_coll = -log₂(Σ pᵢ²)', '较保守\n(★★★★☆)', '基于碰撞概率\n介于min和Shannon之间\n有时更容易估计', '不如min-entropy保守\n不如Shannon熵广泛', '特定统计测试\n补充分析', 'Coron熵测试的理论基础\n黑盒测试套件中的辅助指标', '与碰撞概率的期望值直接相关'],
    ['Guess Work / Work Factor\n(猜测工作因子)', 'w_λ(X) = min{#guesses:\nP(正确)≥λ}', '基于攻击视角\n(实用导向)', '直接量化攻击者工作量\n对安全设计有直接指导意义\n考虑成功概率λ', '计算复杂(对复杂分布)\n依赖于具体攻击模型\n不如熵度量通用', '安全强度论证\n攻击成本分析\n密钥长度选择', '数学背景(4.3.2节)\n用于评估熵的\n实际安全意义', 'w_λ与min-entropy关系:\nH_min ≤ w_λ ≤ H_min - log₂(1-λ)\n(对iid二值随机变量)'],
]
write_table(ws4, headers4, data4, 2)
style_sheet(ws4, cols4, 1)

# ============================================
# Sheet 5: 测试方法汇总
# ============================================
ws5 = wb.create_sheet("测试方法汇总")
ws5.cell(row=1, column=1, value="测试方法汇总 (Testing Methods Summary)").font = title_font
ws5.merge_cells('A1:G1')
cols5 = [18, 14, 20, 12, 22, 14, 18]
headers5 = ['测试类型', '执行时机', '检测目标', '负责方', '关键指标', '适用等级', '相关章节']
data5 = [
    ['在线测试\n(Online Test)', '持续运行期间\n实时执行', '噪声源是否偏离随机模型\n检测统计特性退化', 'RNG自身\n(内部健康测试)', '低误报率(A_real区域)\n高检测率(A_bad区域)\n噪声报警触发概率', 'PTG.2, PTG.3\nNTG.1', '4.5.3节, 5.5节\n在线测试程序实例'],
    ['总失效测试\n(Total Failure Test)', '持续运行期间\n(比在线测试更频繁)', '噪声源完全失效\n(硬件故障、断路等)', 'RNG自身\n(内部健康测试)', '几乎零漏报(必须检测到)\n检测速度: 几乎立即', 'PTG.2, PTG.3\nNTG.1', '4.5.4节\n总失效检测机制'],
    ['启动测试\n(Start-up Test)', '每次上电/复位时\n(输出随机数之前)', '噪声源初始工作状态\n严重的统计缺陷', 'RNG自身\n(内部健康测试)', '启动时通过率\n覆盖总失效和严重弱点', 'PTG.2, PTG.3', '4.5.5节'],
    ['已知答案测试\n(KAT)', '设备启动时\n或定期执行', '算法组件基本功能\n(如AES引擎是否正常)', 'RNG自身\n(内部自检)', '通过/失败', '所有等级\n(超出本文档范围)', '2.1节(范围)\n不属于AIS31评估范围'],
    ['评估者黑盒测试\nT0_rrn', '评估期间\n(一次性，由评估者执行)', '原始随机数的统计质量\n(独立同分布检验)', '评估者\n(外部评估)', '卡方检验p值\nCoron熵测试统计量\n多项统计测试组合', 'PTG.2, PTG.3\n(对原始随机数)', '4.6.2节\nT0_rrn测试套件规范'],
    ['评估者黑盒测试\nT0_irn', '评估期间\n(一次性，由评估者执行)', '内部随机数的统计质量', '评估者\n(外部评估)', '同T0_rrn\n+ 增强的熵测试', 'PTG.2, PTG.3\n(对内部随机数)', '4.6.3节\nT0_irn测试套件规范'],
    ['Coron熵测试', '评估期间\n或作为在线测试的一部分', '每比特熵的统计估计\n基于L比特块的频率', '评估者\n(或RNG自身)', '测试统计量C_L\n与理论阈值比较\nσ_C (标准差)', 'PTG.2, PTG.3\n(辅助评估)', '4.6.1节 (Tables 6-7)\nCoron99论文'],
    ['卡方拟合优度检验\n(χ² GOF Test)', '评估期间\n或作为在线测试的一部分', '原始随机数的分布\n是否符合预期模型\n(如B(1,p)均匀性)', '评估者\n(或RNG自身)', 'χ²统计量\n自由度\np值阈值', 'PTG.2, PTG.3\nNTG.1', '5.5.1节 (Tables 12-13)\n4比特字上的仿真结果'],
]
write_table(ws5, headers5, data5, 2)
style_sheet(ws5, cols5, 1)

# ============================================
# Sheet 6: 噪声源设计示例
# ============================================
ws6 = wb.create_sheet("噪声源设计示例")
ws6.cell(row=1, column=1, value="噪声源设计示例 (Noise Source Design Examples)").font = title_font
ws6.merge_cells('A1:H1')
cols6 = [18, 10, 18, 22, 18, 14, 12, 16]
headers6 = ['设计名称', '类型', '物理原理', '随机模型', '优点', '缺点', '章节', '适用等级']
data6 = [
    ['双噪声二极管\nPTRNG', 'PTRNG\n(物理)', '两个齐纳二极管的\n噪声电压差→放大\n→0-1交叉检测', '时间间隔服从\nGamma分布\n参数通过实测估计', '经典设计，经过充分验证\n熵源质量高\n数学建模完善', '需要精密模拟电路\n器件老化影响\n成本较高', '5.4.2节\n(Killmann设计)', 'PTG.2, PTG.3\n(可作为PTG.3核心)'],
    ['等间隔采样\n事件设计A', 'PTRNG\n(物理)', '等间隔采样物理事件\n中间时间间隔iid\n数字化策略A', 'T_j ~ N(μ, σ²)\nμ=1.0\n仿真: N=10,000,000', '建模简单(正态分布)\n仿真数据充分\n易懂的教学案例', '实际物理实现复杂\niid假设在现实中\n不完全满足', '5.4.3节\n(设计A)', 'PTG.2\n(基础级)'],
    ['等间隔采样\n事件设计B', 'PTRNG\n(物理)', '同设计A\n但数字化策略不同\n(比较不同策略的效果)', 'T_j ~ N(μ, σ²)\n或Gamma分布\n(取决于具体实现)', '展示不同数字化策略\n对随机模型的影响\n教学参考价值高', '同设计A', '5.4.4节\n(设计B)', 'PTG.2\n(基础级,参考)'],
    ['放射性衰变\nPTRNG', 'PTRNG\n(物理/量子)', '放射性同位素\n(如Am-241)的衰变事件\n随机时间间隔', '泊松过程\n时间间隔服从\n指数分布', '物理随机性极好\n(量子效应)\n不可预测性极强', '实现复杂\n放射性材料管理\n输出速率低', '5.4.5节', 'PTG.2, PTG.3\n(高安全性场景)'],
    ['PLL锁相环\n物理噪声源', 'PTRNG\n(物理/电子)', '锁相环(PLL)的\n相位抖动和频率噪声\n数字化采样', 'PLL抖动模型\n频率锁定环路的\n随机相位噪声', '适合集成电路实现\n体积小、成本低\n输出速率高', '建模较复杂\n易受电源噪声\n和温度影响', '5.4.6节', 'PTG.2, PTG.3\n(嵌入式/IC场景)'],
    ['Linux内核\n熵源(NPTRNG)', 'NPTRNG\n(非物理)', '键盘/鼠标中断时间\n磁盘I/O抖动\n网络中断定时\nRDRAND指令等', '无正式随机模型\n(NPTRNG特性)\n基于多种熵源混合\n使用min-entropy下界', '跨平台运行\n无需专用硬件\n成本为零', '熵估计困难(无法精确建模)\n虚拟环境熵源可能不足\n易受攻击者影响', '5.6节\n(Linux RNG)', 'NTG.1\n(可作DRNG种子源)'],
    ['环形振荡器\n(通用类型)', 'PTRNG\n(物理/电子)', '多个环形振荡器的\n相位漂移和热噪声\n→采样→XOR混合', '热噪声的\n随机过程模型\n振荡器间相互\n独立的抖动', '纯数字电路可实现\nFPGA/ASIC友好\n业内广泛使用', '需要多个振荡器\n以消除相关性\n温度补偿需求', '2.3节\n(概述)', 'PTG.2, PTG.3\n(数字IC场景)'],
]
write_table(ws6, headers6, data6, 2)
style_sheet(ws6, cols6, 1)

# ============================================
# Sheet 7: 后处理算法对比
# ============================================
ws7 = wb.create_sheet("后处理算法对比")
ws7.cell(row=1, column=1, value="后处理算法对比 (Post-Processing Algorithms Comparison)").font = title_font
ws7.merge_cells('A1:H1')
cols7 = [16, 10, 18, 18, 14, 10, 14, 16]
headers7 = ['算法名称', '类型', '输入→输出关系', '熵影响', '吞吐率影响', '密码学?', '适用等级', '文档章节']
data7 = [
    ['Von Neumann\n去偏', '算法后处理\n(非密码学)', '成对处理输入比特\n(0,1)→0, (1,0)→1\n(0,0)和(1,1)→丢弃', '消除一阶偏向\n不增加每比特熵\n(仅改善分布均匀性)', '降至 p(1-p) 的\n理论输出率\n一般≤25%', '否', 'PTG.2\n(可选后处理)', '5.1.2节\n经典去偏算法'],
    ['固定压缩率\n(Fixed Compression)', '算法后处理\n(非密码学)', 'n比特输入→m比特输出\n(m < n)\n如8→1, 4→1', '压缩率>1可增加\n每比特熵\n(通过丢弃冗余比特)', '降至 m/n 的\n输入速率', '否', 'PTG.2\n(可选后处理)', '5.1.1节'],
    ['XOR压缩', '算法后处理\n(非密码学)', '多个独立比特\n→XOR→1比特\n或向量XOR', '增加每比特熵\n(如果输入独立)\n消除偏向', '降至 1/k\n(k个输入→1输出)', '否', 'PTG.2\n(可选后处理)', '3.4节\n(简单后处理示例)'],
    ['稀疏化\n(Thinning Out)', '算法后处理\n(非密码学)', '周期性丢弃比特\n(如丢弃每隔一个的字节)', '降低统计依赖性\n不直接增加每比特熵', '降至原始速率的\n(1-d) (d=丢弃比例)', '否', 'PTG.2\n(可选后处理)', '5.1.3节'],
    ['AES-OFB模式', '密码学后处理\n(带记忆)', 'AES在输出反馈模式\n运行: 状态→AES加密\n→输出+新状态', '维持每比特熵水平\n不增加(纯DRNG)\n但提供计算安全性', '接近AES吞吐率\n(约数百MB/s)', '是\n(对称密码)', 'DRG.3, DRG.4\nPTG.3\n(密码学后处理)', '5.2.1节\nAES-OFB评估'],
    ['Hash DRBG\n(SHA-256 based)', '密码学后处理\n(带记忆)', '状态+额外输入→SHA-256\n→内部随机数+新状态\nc_rate可变', '维持/可配置熵水平\n满足DRG.3算法要求', '取决于SHA-256\n实现性能', '是\n(哈希函数)', 'DRG.3, DRG.4\nPTG.3\n(NIST标准)', '5.3.1节\nHash DRBG安全评估'],
    ['SHA-256 OFB\n(示例设计)', '密码学后处理\n(带记忆)', 's∥11∥a→SHA-256→新状态\ns∥00∥a→SHA-256→输出\n256-bit内部状态', '维持256-bit安全水平\n满足c_rate≥1要求', '中等\n(SHA-256软件实现)', '是\n(哈希函数)', 'PTG.3\n(示例教学)', '3.4.4节 (Par.330)\nPTG.3示例计算'],
    ['恒等映射\n(无后处理)', '无后处理', '原始随机数 = 内部随机数\n(直接输出)', '熵不变\n(原始随机数的熵\n=内部随机数的熵)', '100%\n(无开销)', '否', 'PTG.2\n(如原始随机数\n质量足够)', '3.4节\n(允许但需验证\n原始随机数熵充足)'],
]
write_table(ws7, headers7, data7, 2)
style_sheet(ws7, cols7, 1)

# ============================================
# Sheet 8: NIST SP800-90A一致性
# ============================================
ws8 = wb.create_sheet("NIST一致性分析")
ws8.cell(row=1, column=1, value="NIST SP800-90A 与 AIS 31 一致性分析").font = title_font
ws8.merge_cells('A1:F1')
cols8 = [22, 18, 14, 28, 14, 22]
headers8 = ['NIST DRBG / RNG', '目标AIS31等级', '一致性状态', '关键条件', '引用章节', '备注']
data8 = [
    ['Hash DRBG\n(SHA-256/SHA-384/SHA-512)', 'DRG.3', '✓ 满足(经分析验证)', '种子源: TRNG (PTG.2/3/NTG.1)\n请求限制: 符合DRG.3.2\n内部状态: 符合DRG.3.3\nϕ(Hash_df+更新)是单向函数\nψ(Hashgen)是密码学的\n增强后向保密: 满足', '5.3.1节', '开发者可直接引用\n第5.3节的安全分析\n作为认证证据'],
    ['Hash DRBG\n(符合DRG.3条件)', 'DRG.4', '✓ 可满足(需额外条件)', '在DRG.3基础上:\n+ 密码学后处理带记忆\n+ 持续熵输入刷新\n+ 熵源: PTG.2/3或NTG.1', '5.3.1节\n3.3.5节', 'Hash DRBG本身不自动满足DRG.4\n需要设计上增加持续熵刷新机制'],
    ['HMAC DRBG\n(HMAC-SHA-256等)', 'DRG.3', '✓ 满足(经分析验证)', '同Hash DRBG条件\nHMAC作为单向函数\n满足DRG.3算法要求', '5.3节\n(一致性分析)', '与Hash DRBG同级别的安全性\n开发者可引用'],
    ['CTR DRBG\n(AES-CTR based)', 'DRG.3', '✓ 满足(经分析验证)', '同Hash DRBG条件\nAES-CTR满足\n密码学输出函数要求', '5.3节\n(一致性分析)', '与Hash DRBG同级别的安全性\n开发者可引用'],
    ['Linux CRNG\n(ChaCha20-based)', 'DRG.3', '✓ 满足(BSI研究验证)', 'CRNG = ChaCha20-based DRNG\n满足DRG.3算法要求\n熵源: Linux熵池(NTG.1路线)', '5.6节\n(Linux RNG研究)', 'BSI长期研究结论\n在Linux内核5.6+中验证'],
    ['Linux\n/dev/random', 'NTG.1', '✓ 满足(BSI研究验证)', '非物理熵源\nmin-entropy下界量化\n在线测试(阻塞式保证)', '5.6节\n(表16)', '作为NPTRNG认证\n阻塞模式: 保证新鲜熵'],
    ['Linux\n/dev/urandom', 'DRG.3', '✓ 满足(BSI研究验证)', '由CRNG提供DRNG功能\n熵源来自NTG.1熵池\n满足DRG.3所有要求', '5.6节\n(表16)', '非阻塞模式\n不保证新鲜熵\n(但实际安全性高)'],
]
write_table(ws8, headers8, data8, 2)
style_sheet(ws8, cols8, 1)

# ============================================
# Sheet 9: 关键数值参数
# ============================================
ws9 = wb.create_sheet("关键数值参数")
ws9.cell(row=1, column=1, value="关键数值参数 (Key Numerical Parameters)").font = title_font
ws9.merge_cells('A1:E1')
cols9 = [28, 16, 14, 20, 30]
headers9 = ['参数名称', '数值', '单位', '适用范围', '说明']
data9 = [
    ['有效内部状态最小值', '252', 'bits', 'DRG.2, DRG.3, DRG.4', 'DRNG的有效内部状态大小下限\n在攻击者已知大量内部随机数\n的假设下确定'],
    ['初始内部状态 min-entropy最小值', '240', 'bits', 'DRG.2, DRG.3, DRG.4\n(声明min-entropy时)', '种子/重新种子后的初始内部状态\n所需的最小min-entropy量'],
    ['初始内部状态 Shannon熵最小值', '250', 'bits', 'DRG.2, DRG.3, DRG.4\n(声明Shannon熵时)', '种子/重新种子后的初始内部状态\n所需的最小Shannon熵量\n需PTRNG平稳分布'],
    ['最大请求次数 (两次种子间)', '2^48\n(≈2.8×10^14)', '次', 'DRG.2, DRG.3, DRG.4', '两次连续种子/重新种子过程\n之间允许的最大请求数'],
    ['单次请求最大长度', '2^19\n(524,288)', 'bits\n(64 KB)', 'DRG.2, DRG.3', '单次请求的最大比特长度\n(DRG.4可能更严格)'],
    ['PTG.2 Shannon熵每比特', '0.9998', 'per bit', 'PTG.2\n(声明Shannon熵时)', '每内部随机数比特的平均Shannon熵\n下限。接近1但允许微小缺陷'],
    ['PTG.2 Min-entropy每比特', '0.98', 'per bit', 'PTG.2\n(声明min-entropy时)', '每内部随机数比特的平均min-entropy\n下限。保守但更可靠'],
    ['PTG.3 最小压缩率 c_rate', '1', 'ratio\n(n/m)', 'PTG.3', '密码学后处理的输入比特数/输出比特数\n必须≥1。要增加熵需>1'],
    ['A4页面宽度 (默认)', '11,906', 'DXA\n(1/1440 inch)', '通用 (文档生成参考)', 'docx-js默认页面宽度\n1 DXA = 1/1440 inch'],
    ['US Letter页面宽度', '12,240', 'DXA', '通用 (文档生成参考)', '美国标准纸张宽度\n(8.5 inches × 1440)'],
    ['A4内容宽度 (1"页边距)', '9,026', 'DXA', '通用 (文档生成参考)', '页面宽度 - 左右页边距\n= 11906 - 2880 = 9026 DXA'],
    ['US Letter内容宽度 (1"页边距)', '9,360', 'DXA', '通用 (文档生成参考)', '页面宽度 - 左右页边距\n= 12240 - 2880 = 9360 DXA'],
    ['最大种子树高度', '限制 (≥1)', '层', 'DRG.3\n(种子树结构)', '根DRNG(高度0) + 子DRNG\n允许的最大层级数。过高会增加\n安全风险'],
    ['在线测试 Chi²自由度', '取决于字长w\n(如w=4→15df)', 'df', 'PTG.2, PTG.3\nNTG.1', '卡方拟合优度检验的自由度\n= 2^w - 1 (对w比特字)'],
    ['Coron熵测试 L参数', 'L=8\n(典型值)', 'bits\n(块大小)', 'PTG.2, PTG.3\n(辅助评估)', 'Coron测试中将原始比特序列\n划分为L比特块。L=8为典型配置\n但也支持其他值'],
    ['AIS 31 文档页数', '239', '页', '全部', 'AIS 31 v2.35 DRAFT\n(2022年9月版)\n完整技术参考文档的页数'],
    ['噪音报警评估准则数', '3', '个', 'PTG.2, PTG.3, NTG.1\n(在线测试)', 'ER1a/b: 样本计数阈值\nER2: 连续测试统计量阈值\nER3: 长期统计指标'],
]
write_table(ws9, headers9, data9, 2)
style_sheet(ws9, cols9, 1)

# ============================================
# Save
# ============================================
wb.save('D:\\ClaudeCodeTest\\test5\\AIS31_参考表格.xlsx')
print("Excel file saved successfully!")
print(f"Sheets: {wb.sheetnames}")
